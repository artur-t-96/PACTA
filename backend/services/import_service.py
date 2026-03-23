"""
Import service — import historical contracts from Marta's Excel
"""
import os
import re
from datetime import datetime
from typing import Optional
import openpyxl
from sqlalchemy.orm import Session
from database import Contract, ContractAuditLog
import json


def parse_date(val) -> Optional[str]:
    """Parse date from various Excel formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() in ("none", "nieokreślony", "brak"):
        return None
    # Try common formats
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s[:10] if len(s) >= 10 else s


def parse_name(full_name: str) -> tuple:
    """Parse 'Nazwisko Imię' → (imie, nazwisko)."""
    parts = full_name.strip().split()
    if len(parts) >= 2:
        nazwisko = parts[0]
        imie = " ".join(parts[1:])
        return imie, nazwisko
    return full_name, ""


def normalize_contract_number(num) -> str:
    """Normalize contract number."""
    if num is None:
        return ""
    s = str(num).strip()
    # Remove .0 from float parsing
    if s.endswith(".0"):
        s = s[:-2]
    return s


def determine_status(end_date_str: Optional[str], uwagi: str = "") -> str:
    """Determine contract status from end date and notes."""
    uwagi_lower = (uwagi or "").lower()
    
    if any(w in uwagi_lower for w in ["wypowiedzenie", "zakończona", "rozwiązana", "rezygnacja"]):
        return "zakonczona"
    if any(w in uwagi_lower for w in ["anulowana", "odwołana"]):
        return "anulowana"
    
    if end_date_str and end_date_str != "nieokreślony":
        try:
            end_date = datetime.strptime(end_date_str, "%Y-%m-%d")
            if end_date < datetime.now():
                return "zakonczona"
            return "aktywna"
        except ValueError:
            pass
    
    return "aktywna"  # Default: if no end date → still active


def import_all_sheets(file_path: str, db: Session) -> dict:
    """Import from all sheets."""
    total = {"imported": 0, "skipped": 0, "errors": [], "duplicates": 0}
    for sheet in ["Umowy B2B", "Bez dzialalności "]:
        try:
            result = import_from_excel(file_path, db, sheet)
            total["imported"] += result["imported"]
            total["duplicates"] += result["duplicates"]
            total["errors"] += result["errors"]
        except Exception as e:
            total["errors"].append(f"Sheet '{sheet}': {str(e)}")
    return total


def import_from_excel(file_path: str, db: Session, sheet_name: str = "Umowy B2B") -> dict:
    """
    Import historical contracts from Marta's Excel.
    Returns import stats.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    
    stats = {"imported": 0, "skipped": 0, "errors": [], "duplicates": 0}
    
    for row in range(2, ws.max_row + 1):
        full_name = ws.cell(row, 1).value
        if not full_name:
            break
        
        full_name = str(full_name).strip()
        imie, nazwisko = parse_name(full_name)
        numer = normalize_contract_number(ws.cell(row, 2).value)
        klient = str(ws.cell(row, 3).value or "Brak").strip()
        stanowisko = str(ws.cell(row, 4).value or "").strip()
        rodzaj = str(ws.cell(row, 5).value or "B2B").strip()
        data_podpisania = parse_date(ws.cell(row, 6).value)
        data_startu = parse_date(ws.cell(row, 7).value)
        data_zakonczenia = parse_date(ws.cell(row, 8).value)
        okres_lojalnosci = str(ws.cell(row, 9).value or "").strip()
        rekruter = str(ws.cell(row, 10).value or "").strip()
        uwagi = str(ws.cell(row, 13).value or "").strip()
        zmiany = str(ws.cell(row, 14).value or "").strip()
        
        # Normalize client names
        klient_map = {
            "BNP": "BNP Paribas Bank Polska S.A.",
            "Nordea": "Nordea Bank Abp S.A. Oddział w Polsce",
            "B2B.NET": "B2B.net S.A.",
            "PKO BP": "PKO Bank Polski S.A.",
            "PKO": "PKO Bank Polski S.A.",
            "Alior": "Alior Bank S.A.",
        }
        klient_norm = klient_map.get(klient, klient)
        
        # Build contract number
        if numer:
            # Historical numbers: "361" → "H-361"
            contract_number = f"H-{numer}" if not "/" in str(numer) else str(numer)
        else:
            contract_number = f"H-{row}"
        
        # Check duplicates
        existing = db.query(Contract).filter(Contract.number == contract_number).first()
        if existing:
            stats["duplicates"] += 1
            continue
        
        # Determine status
        status = determine_status(
            data_zakonczenia if data_zakonczenia != "nieokreślony" else None,
            uwagi
        )
        
        # Create contract
        try:
            contract = Contract(
                number=contract_number,
                contractor_name=f"{imie} {nazwisko}",
                contractor_firstname=imie,
                contractor_lastname=nazwisko,
                contractor_company="",
                contractor_nip="",
                contractor_regon="",
                contractor_address="",
                contractor_email="",
                contractor_phone="",
                client=klient_norm,
                role=stanowisko,
                rate=0.0,  # Not in Excel
                it_area=_guess_it_area(stanowisko),
                it_area_raw=stanowisko,
                project_description=uwagi if uwagi != "None" else "",
                client_city="",
                start_date=data_startu or data_podpisania or "",
                status=status,
                file_path="",
            )
            db.add(contract)
            db.flush()
            
            # Audit log
            details = {"source": "excel_import", "row": row, "rodzaj": rodzaj}
            if rekruter and rekruter != "None":
                details["rekruter"] = rekruter
            if zmiany and zmiany != "None":
                details["zmiany_historyczne"] = zmiany[:500]
            if okres_lojalnosci and okres_lojalnosci != "None":
                details["okres_lojalnosci"] = okres_lojalnosci
            
            db.add(ContractAuditLog(
                contract_id=contract.id,
                action="imported",
                details=json.dumps(details, ensure_ascii=False),
                user="excel_import",
            ))
            
            stats["imported"] += 1
        except Exception as e:
            stats["errors"].append(f"Row {row} ({full_name}): {str(e)}")
    
    db.commit()
    return stats


def _guess_it_area(stanowisko: str) -> str:
    """Guess IT area from role name."""
    s = stanowisko.lower()
    
    if any(w in s for w in ["test", "qa", "quality", "tester"]):
        return "QA"
    if any(w in s for w in ["devops", "ci/cd", "sre", "infrastructure", "cloud engineer"]):
        return "DevOps"
    if any(w in s for w in ["data", "analityk", "bi", "etl", "warehouse"]):
        return "Data & Analytics"
    if any(w in s for w in ["cloud", "aws", "azure", "architect"]):
        return "Cloud Solutions"
    if any(w in s for w in ["automat", "selenium", "cypress"]):
        return "Test Automation"
    if any(w in s for w in ["develop", "programist", "java", "python", ".net", "react", "frontend", "backend", "fullstack"]):
        return "Software Development"
    if any(w in s for w in ["admin", "support", "helpdesk"]):
        return "IT Support"
    if any(w in s for w in ["project", "scrum", "agile", "pm"]):
        return "Project Management"
    if any(w in s for w in ["security", "cyber", "soc", "pentester"]):
        return "Cybersecurity"
    
    return "Other"
