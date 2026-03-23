"""
Exports router — CSV, XLSX, ZIP, per-client exports
"""
import os
import io
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from database import SessionLocal, Contract

router = APIRouter()


@router.get("/csv")
def export_csv(status: str = None):
    """Export contracts as CSV."""
    import csv
    db = SessionLocal()
    query = db.query(Contract)
    if status: query = query.filter(Contract.status == status)
    contracts = query.order_by(Contract.created_at.desc()).all()
    
    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(["Numer", "Kontrahent", "Firma", "NIP", "Klient", "Rola",
                      "Stawka PLN/h", "Obszar IT", "Data startu", "Status", "Data utworzenia"])
    for c in contracts:
        writer.writerow([c.number, c.contractor_name, c.contractor_company, c.contractor_nip,
                         c.client, c.role, c.rate, c.it_area, c.start_date, c.status, c.created_at])
    db.close()
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": "attachment; filename=umowy_b2b.csv"})


@router.get("/xlsx")
def export_xlsx(status: str = None):
    """Export as colored Excel."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    
    db = SessionLocal()
    query = db.query(Contract)
    if status: query = query.filter(Contract.status == status)
    contracts = query.order_by(Contract.status, Contract.created_at.desc()).all()
    db.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Umowy B2B"
    
    headers = ["Nr umowy", "Kontrahent", "Firma", "NIP", "Klient", "Rola",
               "Stawka PLN/h", "Obszar IT", "Data startu", "Status", "Opis", "Data utworzenia"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E40AF")
        cell.alignment = Alignment(horizontal="center")
    
    status_colors = {"aktywna": "D1FAE5", "draft": "FEF3C7", "do_podpisu": "EDE9FE",
                     "zakonczona": "F3F4F6", "anulowana": "FEE2E2"}
    
    for row, c in enumerate(contracts, 2):
        fill = PatternFill("solid", fgColor=status_colors.get(c.status, "FFFFFF"))
        values = [c.number, c.contractor_name, c.contractor_company, c.contractor_nip,
                  c.client, c.role, c.rate, c.it_area, c.start_date, c.status,
                  (c.project_description or "")[:100], str(c.created_at)[:10]]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row, col, val)
            cell.fill = fill
    
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=umowy_b2b.xlsx"})


@router.get("/zip")
def export_zip(ids: str = ""):
    """Download multiple DOCXs as ZIP."""
    import zipfile
    db = SessionLocal()
    if ids:
        id_list = [int(x) for x in ids.split(",") if x.strip().isdigit()]
        contracts = db.query(Contract).filter(Contract.id.in_(id_list)).all()
    else:
        contracts = db.query(Contract).filter(
            Contract.file_path != "", Contract.file_path != None,
            ~Contract.number.startswith("H-")
        ).limit(50).all()
    db.close()
    
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
        for c in contracts:
            if c.file_path and os.path.exists(c.file_path):
                zf.write(c.file_path, os.path.basename(c.file_path))
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="application/zip",
                             headers={"Content-Disposition": "attachment; filename=umowy_b2b.zip"})


@router.get("/client/{client_name}")
def export_client(client_name: str):
    """Export all contracts for a specific client."""
    import urllib.parse, openpyxl
    from openpyxl.styles import Font, PatternFill
    
    name = urllib.parse.unquote(client_name)
    db = SessionLocal()
    contracts = db.query(Contract).filter(Contract.client.ilike(f"%{name}%")).order_by(
        Contract.status, Contract.created_at.desc()).all()
    db.close()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = name[:31]
    headers = ["Nr", "Kontrahent", "Firma", "NIP", "Rola", "Stawka", "Obszar IT", "Start", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E40AF")
    for c in contracts:
        ws.append([c.number, c.contractor_name, c.contractor_company, c.contractor_nip,
                   c.role, c.rate, c.it_area, c.start_date, c.status])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    safe = name.replace(" ", "_").replace("/", "_")[:30]
    return StreamingResponse(iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={safe}.xlsx"})
